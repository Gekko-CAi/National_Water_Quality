# -*- coding: utf-8 -*-
"""
国家水质自动综合监管平台 - 数据抓取脚本
从 https://szzdjc.cnemc.cn:8070/GJZ/Business/Publish/Main.html 抓取实时水质监测数据
每4小时运行一次，数据保存到 Excel 文件中，每个文件保存5-10天数据
"""

import os
import sys
import re
import json
import time
import glob
from datetime import datetime, timedelta
from pathlib import Path

import requests
import urllib3
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 禁用 SSL 警告
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ======================== 配置 ========================
API_URL = "https://szzdjc.cnemc.cn:8070/GJZ/Ajax/Publish.ashx"
REFERER_URL = "https://szzdjc.cnemc.cn:8070/GJZ/Business/Publish/RealDatas.html"
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "数据文件")
MAX_DAYS_PER_FILE = 10  # 每个 Excel 文件最多保存天数
PAGE_SIZE = 2000  # 每页记录数（设大一点减少请求次数）
REQUEST_TIMEOUT = 60  # 请求超时秒数
PAGE_DELAY = 2  # 每页请求间隔秒数（避免请求过快）

# Git 自动提交配置
GIT_AUTO_COMMIT = True  # 是否自动 git commit
GIT_AUTO_PUSH = True    # 是否自动 git push（需先配置 remote）
PROJECT_DIR = os.path.dirname(os.path.abspath(__file__))

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Content-Type": "application/x-www-form-urlencoded",
    "Referer": REFERER_URL,
    "X-Requested-With": "XMLHttpRequest",
    "Origin": "https://szzdjc.cnemc.cn:8070",
}

# 水质类别映射
WATER_QUALITY_MAP = {
    "1": "I类", "2": "II类", "3": "III类",
    "4": "IV类", "5": "V类", "6": "劣V类", "7": "未监测"
}

# 水质类别颜色 (openpyxl PatternFill)
QUALITY_COLORS = {
    "I类": PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid"),
    "II类": PatternFill(start_color="00CCFF", end_color="00CCFF", fill_type="solid"),
    "III类": PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid"),
    "IV类": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
    "V类": PatternFill(start_color="FF9B00", end_color="FF9B00", fill_type="solid"),
    "劣V类": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
    "未监测": PatternFill(start_color="ABABAB", end_color="ABABAB", fill_type="solid"),
}


# ======================== 数据抓取 ========================
def parse_cell_value(html_str):
    """
    解析单元格 HTML 值，提取原始值和显示值。
    格式: <span title='原始值：23.88'>23.9</span> 或 -- 或 纯文本
    返回: (原始值, 显示值)
    """
    if not html_str or html_str == "--":
        return ("--", "--")

    # 尝试提取原始值
    original_match = re.search(r"原始值[：:]\s*([\d.]+)", html_str)
    # 尝试提取显示值（span 内的文本）
    display_match = re.search(r">([^<]+)<", html_str)

    original_val = original_match.group(1) if original_match else html_str
    display_val = display_match.group(1) if display_match else html_str

    # 清理 HTML 标签
    if "<" in original_val:
        original_val = re.sub(r"<[^>]+>", "", original_val)
    if "<" in display_val:
        display_val = re.sub(r"<[^>]+>", "", display_val)

    return (original_val.strip(), display_val.strip())


def parse_thead(thead_list):
    """解析表头，提取干净的列名"""
    clean_headers = []
    for h in thead_list:
        # 移除 HTML 标签
        clean = re.sub(r"<[^>]+>", "", h)
        # 移除多余空白
        clean = re.sub(r"\s+", " ", clean).strip()
        clean_headers.append(clean)
    return clean_headers


def fetch_page(page_index, page_size=PAGE_SIZE, max_retries=3):
    """抓取单页数据"""
    data = {
        "action": "getRealDatas",
        "AreaID": "",
        "RiverID": "",
        "MNName": "",
        "PageIndex": str(page_index),
        "PageSize": str(page_size),
    }

    for attempt in range(max_retries):
        try:
            response = requests.post(
                API_URL, headers=HEADERS, data=data,
                verify=False, timeout=REQUEST_TIMEOUT
            )
            result = response.json()
            if result.get("result") and result["result"] != 0:
                return result
            else:
                print(f"  [警告] 第{page_index}页返回空数据")
                return None
        except requests.exceptions.Timeout:
            print(f"  [重试 {attempt+1}/{max_retries}] 第{page_index}页请求超时")
            time.sleep(5)
        except Exception as e:
            print(f"  [重试 {attempt+1}/{max_retries}] 第{page_index}页请求异常: {e}")
            time.sleep(5)

    print(f"  [错误] 第{page_index}页抓取失败（已重试{max_retries}次）")
    return None


def fetch_all_data():
    """抓取所有页的数据"""
    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 开始抓取水质监测数据...")

    # 第一页
    first_page = fetch_page(1)
    if not first_page:
        print("[错误] 第一页数据抓取失败，无法继续")
        return None

    total_pages = first_page.get("total", 1)
    total_records = first_page.get("records", 0)
    thead = first_page.get("thead", [])
    all_tbody = first_page.get("tbody", [])

    clean_headers = parse_thead(thead)
    print(f"  总记录数: {total_records}, 总页数: {total_pages}")
    print(f"  表头列: {clean_headers}")

    # 抓取剩余页
    for page_idx in range(2, total_pages + 1):
        print(f"  正在抓取第 {page_idx}/{total_pages} 页...")
        time.sleep(PAGE_DELAY)
        page_data = fetch_page(page_idx)
        if page_data and page_data.get("tbody"):
            all_tbody.extend(page_data["tbody"])
        else:
            print(f"  [警告] 第{page_idx}页无数据，跳过")

    print(f"  抓取完成，共获取 {len(all_tbody)} 条记录")
    return {"headers": clean_headers, "data": all_tbody}


# ======================== 数据处理 ========================
def process_data(raw_data):
    """
    将原始数据解析为结构化记录列表。
    每条记录是一个 dict，包含所有字段。
    """
    headers = raw_data["headers"]
    rows = raw_data["data"]
    records = []

    # 解析监测时间，添加年份
    current_year = datetime.now().year

    for row in rows:
        record = {}
        for i, val in enumerate(row):
            if i < len(headers):
                col_name = headers[i]
            else:
                col_name = f"列{i+1}"

            if i == 4:  # 水质类别
                quality_level = WATER_QUALITY_MAP.get(str(val), val)
                record[col_name] = quality_level
            elif i >= 5:  # 监测指标值（含 HTML）
                original, display = parse_cell_value(val)
                record[col_name] = original  # 使用原始值
            elif i == 3:  # 监测时间
                # 格式: "07-01 20:00" -> 添加年份
                time_str = str(val).strip()
                if time_str and time_str != "--":
                    record[col_name] = f"{current_year}-{time_str}"
                else:
                    record[col_name] = time_str
            else:
                record[col_name] = str(val).strip() if val else ""

        # 添加抓取时间戳
        record["抓取时间"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        records.append(record)

    return records, headers


# ======================== Excel 管理 ========================
def get_excel_files():
    """获取输出目录中所有 Excel 文件，按修改时间排序"""
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
        return []

    files = glob.glob(os.path.join(OUTPUT_DIR, "国控断面水质_*.xlsx"))
    files.sort(key=os.path.getmtime, reverse=True)
    return files


def get_file_date_range(filepath):
    """从 Excel 文件名解析日期范围"""
    basename = os.path.basename(filepath)
    # 文件名格式: 国控断面水质_20250701_20250705.xlsx
    match = re.match(r"国控断面水质_(\d{8})_(\d{8})\.xlsx", basename)
    if match:
        start_str, end_str = match.groups()
        try:
            start_date = datetime.strptime(start_str, "%Y%m%d")
            end_date = datetime.strptime(end_str, "%Y%m%d")
            return start_date, end_date
        except ValueError:
            pass
    return None, None


def find_or_create_excel(records, headers):
    """
    查找当前应使用的 Excel 文件，或创建新文件。
    规则:
    - 如果最新文件的日期跨度 < MAX_DAYS_PER_FILE 天，追加数据
    - 否则创建新文件
    """
    files = get_excel_files()
    today = datetime.now().strftime("%Y%m%d")

    if files:
        latest_file = files[0]
        start_date, end_date = get_file_date_range(latest_file)

        if start_date and end_date:
            days_span = (end_date - start_date).days + 1
            if days_span < MAX_DAYS_PER_FILE:
                # 追加到现有文件
                print(f"  追加数据到现有文件: {os.path.basename(latest_file)} (已跨{days_span}天)")
                append_to_excel(latest_file, records, headers)
                # 更新文件名中的结束日期
                new_end_date = datetime.now().strftime("%Y%m%d")
                if new_end_date != end_date.strftime("%Y%m%d"):
                    new_name = os.path.join(OUTPUT_DIR,
                        f"国控断面水质_{start_date.strftime('%Y%m%d')}_{new_end_date}.xlsx")
                    os.rename(latest_file, new_name)
                    print(f"  文件重命名为: {os.path.basename(new_name)}")
                    return new_name
                return latest_file

    # 创建新文件
    filename = f"国控断面水质_{today}_{today}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)
    print(f"  创建新文件: {filename}")
    create_new_excel(filepath, records, headers)
    return filepath


def create_new_excel(filepath, records, headers):
    """创建新的 Excel 文件并写入数据"""
    wb = Workbook()
    ws = wb.active
    ws.title = "水质监测数据"

    # 添加表头
    all_headers = list(headers) + ["抓取时间"]
    write_header_row(ws, all_headers)

    # 写入数据
    for row_idx, record in enumerate(records, start=2):
        write_data_row(ws, row_idx, record, all_headers)

    # 设置样式
    apply_styles(ws, len(all_headers), len(records) + 1)

    # 冻结首行
    ws.freeze_panes = "A2"

    wb.save(filepath)
    print(f"  数据已保存: {filepath} (共{len(records)}条记录)")


def append_to_excel(filepath, records, headers):
    """追加数据到现有 Excel 文件"""
    wb = load_workbook(filepath)
    ws = wb.active

    # 找到最后一行
    max_row = ws.max_row
    all_headers = list(headers) + ["抓取时间"]

    # 检查表头是否匹配
    existing_headers = [cell.value for cell in ws[1]]
    if existing_headers != all_headers:
        print(f"  [警告] 表头不匹配，将使用新表头")
        # 如果表头不匹配，需要处理
        # 这里简单处理：如果新表头更多，补充列
        if len(all_headers) > len(existing_headers):
            for i in range(len(existing_headers), len(all_headers)):
                ws.cell(row=1, column=i+1, value=all_headers[i])

    # 写入数据
    for row_idx, record in enumerate(records, start=max_row + 1):
        write_data_row(ws, row_idx, record, all_headers)

    wb.save(filepath)
    total = ws.max_row - 1
    print(f"  数据已追加: {filepath} (当前共{total}条记录)")


def write_header_row(ws, headers):
    """写入表头行"""
    header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def write_data_row(ws, row_idx, record, headers):
    """写入一行数据"""
    data_font = Font(name="微软雅黑", size=9)
    data_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_idx, header in enumerate(headers, start=1):
        value = record.get(header, "")
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = data_font
        cell.alignment = data_align
        cell.border = thin_border

        # 水质类别着色
        if header == "水质类别" and value in QUALITY_COLORS:
            cell.fill = QUALITY_COLORS[value]


def apply_styles(ws, num_cols, num_rows):
    """设置列宽等样式"""
    col_widths = {
        "省份": 10, "流域": 12, "断面名称": 16, "监测时间": 16,
        "水质类别": 10, "抓取时间": 20,
    }
    for col_idx in range(1, num_cols + 1):
        header_val = ws.cell(row=1, column=col_idx).value
        width = 12  # 默认宽度
        for key, w in col_widths.items():
            if header_val and key in str(header_val):
                width = w
                break
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 35


# ======================== Git 自动提交 ========================
def git_commit_and_push(filepath):
    """自动将数据文件提交到 Git 仓库，可选推送到远程"""
    if not GIT_AUTO_COMMIT:
        return

    import subprocess

    def run_git(cmd):
        try:
            result = subprocess.run(
                cmd, cwd=PROJECT_DIR, capture_output=True, text=True, timeout=30
            )
            return result.returncode == 0, result.stdout.strip(), result.stderr.strip()
        except Exception as e:
            return False, "", str(e)

    # 添加文件
    rel_path = os.path.relpath(filepath, PROJECT_DIR)
    success, out, err = run_git(["git", "add", rel_path])
    if not success and err:
        print(f"  [Git] add 失败: {err}")
        return

    # 提交
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    msg = f"数据更新: {timestamp} — {os.path.basename(filepath)}"
    success, out, err = run_git(["git", "commit", "-m", msg])
    if success:
        # 只显示摘要行
        summary = out.split("\n")[0] if out else "ok"
        print(f"  [Git] {summary}")
    else:
        # "nothing to commit" 不是错误
        if "nothing to commit" in err:
            print(f"  [Git] 数据无变化，跳过提交")
        else:
            print(f"  [Git] commit 失败: {err}")

    # 推送
    if GIT_AUTO_PUSH:
        success, out, err = run_git(["git", "push"])
        if success:
            print(f"  [Git] push 成功")
        else:
            print(f"  [Git] push 失败: {err}")
            print(f"  [Git] 提示: 先配置 remote: git remote add origin <仓库地址>")


# ======================== 主函数 ========================
def main():
    """主函数：抓取数据并保存到 Excel"""
    # 确保输出目录存在
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)

    # 抓取数据
    raw_data = fetch_all_data()
    if not raw_data or not raw_data.get("data"):
        print("[错误] 未获取到数据，程序退出")
        return False

    # 处理数据
    records, headers = process_data(raw_data)
    if not records:
        print("[错误] 数据处理后为空，程序退出")
        return False

    # 保存到 Excel
    print(f"\n正在保存数据到 Excel...")
    filepath = find_or_create_excel(records, headers)

    # Git 自动提交
    git_commit_and_push(filepath)

    print(f"\n{'='*60}")
    print(f"抓取完成！")
    print(f"  数据文件: {filepath}")
    print(f"  本次记录数: {len(records)}")
    print(f"  抓取时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*60}")
    return True


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
