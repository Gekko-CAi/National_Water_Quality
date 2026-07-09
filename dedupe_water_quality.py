# -*- coding: utf-8 -*-
"""对抓取的 Excel 数据按 (断面名称, 监测时间) 去重，自动检测最新文件"""
import os
import glob
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "数据文件")
KEY_COLS = ("断面名称", "监测时间")
KEEP_TS_COL = "抓取时间"


def find_latest_file():
    """自动查找最新的水质监测数据 Excel 文件"""
    pattern = os.path.join(DATA_DIR, "国控断面水质_*.xlsx")
    files = glob.glob(pattern)
    if not files:
        print(f"[错误] 在 {DATA_DIR} 中未找到数据文件")
        return None
    files.sort(key=os.path.getmtime, reverse=True)
    return files[0]

def col_index(header, name):
    for i, h in enumerate(header, start=1):
        if h == name:
            return i
    return None

def main():
    filepath = find_latest_file()
    if not filepath:
        return
    print(f"处理文件: {os.path.basename(filepath)}")
    wb = load_workbook(filepath)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    key_idx = [col_index(headers, k) for k in KEY_COLS]
    ts_idx = col_index(headers, KEEP_TS_COL)

    if any(i is None for i in key_idx):
        print(f"[错误] 找不到关键列: {KEY_COLS}, 表头={headers}")
        return
    if ts_idx is None:
        print(f"[警告] 找不到 {KEEP_TS_COL} 列，将按出现顺序保留首条")

    # 1) 统计去重前的重复情况
    key_to_rows = {}
    rows = list(ws.iter_rows(min_row=2, values_only=False))
    for row in rows:
        key = tuple(row[i - 1].value for i in key_idx)
        key_to_rows.setdefault(key, []).append(row)

    duplicate_groups = {k: v for k, v in key_to_rows.items() if len(v) > 1}
    dup_rows_count = sum(len(v) - 1 for v in duplicate_groups.values())
    print(f"去重前总行数: {len(rows)}")
    print(f"出现重复的 (断面, 监测时间) 组合数: {len(duplicate_groups)}")
    print(f"可移除的重复行数: {dup_rows_count}")

    # 2) 决定保留行：每组保留抓取时间最晚的那条
    rows_to_keep = set()
    for key, group in key_to_rows.items():
        if len(group) == 1:
            rows_to_keep.add(id(group[0]))
            continue
        def sort_key(r):
            return (r[ts_idx - 1].value or "")
        group_sorted = sorted(group, key=sort_key, reverse=True)
        rows_to_keep.add(id(group_sorted[0]))

    if dup_rows_count == 0:
        print("无需去重")
        wb.close()
        return

    # 3) 重写工作表
    new_rows = [r for r in rows if id(r) in rows_to_keep]
    print(f"去重后保留行数: {len(new_rows)}")

    # 删除原数据行（保留表头）
    ws.delete_rows(2, ws.max_row)

    # 写入新数据
    data_font = Font(name="微软雅黑", size=9)
    data_align = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for r_idx, row_cells in enumerate(new_rows, start=2):
        for c_idx, cell in enumerate(row_cells, start=1):
            new_cell = ws.cell(row=r_idx, column=c_idx, value=cell.value)
            new_cell.font = data_font
            new_cell.alignment = data_align
            new_cell.border = thin

    wb.save(filepath)
    print(f"已保存去重结果到: {filepath}")

if __name__ == "__main__":
    main()
